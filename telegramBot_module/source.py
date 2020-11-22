import csv
import glob
import os
import re
from io import StringIO

import requests
import shutil
from functools import reduce
from operator import concat
import pandas as pd

from openpyxl import load_workbook, Workbook
import schedule

from modules.doeparser import controller
from modules.doeparser import permanent
from modules.core.permanent import DATABASE_FOLDER, DATABASE_NAME
from modules.core.source import bot
from modules.schedule.classes import Group
from modules.schedule.permanent import REGISTERED_COURSES
from modules.admin.permanent import SUPERADMIN_LIST

"""
Module automatically parse schedule from google sheet and modify database

Author: @Nmikriukov
"""


def attach_autoparser_module():
    def get_value(ws, row, col):
        """
        Get value of specific cell from worksheet
        If cell is part of merged cell, return value from top left single cell, where text is stored

        :param ws: worksheet from openpyxl workbook
        :param row: integer
        :param col: integer
        :return: text
        """
        # check if cell is merged
        for borders in ws.merged_cells.ranges:
            if borders.min_col <= col <= borders.max_col and borders.min_row <= row <= borders.max_row:
                return ws.cell(borders.min_row, borders.min_col).value
        # not merged cell
        return ws.cell(row, col).value

    def parse_cell(ws, row, col):
        """
        Get lesson, teacher and room from specific cell

        :param ws: worksheet from openpyxl workbook
        :param row: integer
        :param col: integer
        :return: text, text, text | if correct cell with data
                 None, None, None | if empty cell
                 -1, None, None   | if unknown data in cell
        """
        lesson = get_value(ws, row, col)
        teacher = get_value(ws, row + 1, col)
        room = get_value(ws, row + 2, col)
        if not lesson or len(lesson) < 2:
            return None, None, None  # empty cell

        if "reserve" in lesson.lower():
            return None, None, None

        # remove () brackets and strip
        if lesson and "English" not in lesson:
            lesson = re.sub(r"\(.+\)", "", lesson).strip()
        if teacher:
            teacher = re.sub(r"\(.+\)", "", teacher).strip()
        if isinstance(room, str):
            room = re.sub(r"\(.+\)", "", room).strip()

        if not teacher:  # unknown teacher
            teacher = '?'
        try:
            if "English" not in lesson:
                room = int(room)
            else:
                map(int, room.split("/"))
        except (TypeError, ValueError, AttributeError):
            room = -1

        return lesson, teacher, room

    # specific Exception for download error
    class ScheduleDownloadError(Exception):
        pass

    def parse_new_timetable():
        """
        Download xlsx schedule from link and parse all lessons
        Stores two previous versions of databases and xlsx files
        """
        try:
            # move previous backups
            shutil.move(f"{DATABASE_FOLDER}/{permanent.DATABASE_BACKUP_1}",
                        f"{DATABASE_FOLDER}/{permanent.DATABASE_BACKUP_2}")
            shutil.move(f"{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_1}",
                        f"{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_2}")
        except FileNotFoundError:
            pass
        # compare with previous version of database if such is found
        compare_with_prev = permanent.ADMIN_NOTIFY_TABLE_CHANGES
        try:
            # make new backup
            shutil.copy(f"{DATABASE_FOLDER}/{DATABASE_NAME}",
                        f"{DATABASE_FOLDER}/{permanent.DATABASE_BACKUP_1}")
            shutil.move(f"{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}",
                        f"{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_1}")
        except FileNotFoundError:
            compare_with_prev = False

        # download new schedule from google sheet
        new_schedule = requests.get(permanent.SCHEDULE_DOWNLOAD_LINK)
        # with open(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME_CSV}', 'wb') as f:
        #     f.write(new_schedule.content)
        pd.read_csv(StringIO(new_schedule.text), sep=",").to_excel(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}',
                                                                   header=True, index=False)

        # convert here
        # for csvfile in glob.glob(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME_CSV}'):
        #     workbook = Workbook(write_only=False)
        #     worksheet = workbook.create_sheet("a")
        #     with open(csvfile, 'rt', encoding='utf8') as f:
        #         reader = csv.reader(f)
        #         for r, row in enumerate(reader):
        #             for c, col in enumerate(row):
        #                 worksheet.cell(r + 1, c + 1).value = col
        #     workbook.save(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}')

        try:
            # check download is ok
            schedule_size = shutil.os.path.getsize(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}')
            if schedule_size < permanent.SCHEDULE_MIN_SIZE_BYTES:
                raise ScheduleDownloadError
        except (FileNotFoundError, ScheduleDownloadError):
            # send error notification to admins
            for admin in SUPERADMIN_LIST:
                bot.send_message(admin, permanent.MESSAGE_ERROR_NOTIFY)
            return

        # delete all lessons because new ones will be parsed
        controller.delete_all_lessons()

        # open workbook
        wb = load_workbook(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}')

        sheet_index = 0  # default sheet index in timetable
        # find sheet for bachelors and masters
        for i, name in enumerate(wb.sheetnames):
            if "BS" in name:
                sheet_index = i
        ws = wb[wb.sheetnames[sheet_index]]

        # open workbook from backup
        wb_old, ws_old = None, None
        if compare_with_prev:
            wb_old = load_workbook(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_1}')
            ws_old = wb_old[wb_old.sheetnames[sheet_index]]

        # iterate over each cell
        col = 2
        all_course_groups = reduce(concat, [REGISTERED_COURSES[x] for x in REGISTERED_COURSES])
        memorized_lessons = []  # hack to not insert lesson twice because of multiple columns for one group
        while col <= permanent.SCHEDULE_LAST_COLUMN:
            course_group = get_value(ws, 2, col)
            if course_group not in all_course_groups:
                for admin in SUPERADMIN_LIST:
                    bot.send_message(admin, f"{permanent.MESSAGE_ERROR_UNKNOWN_GROUP}: {course_group}")
                col += 1
                continue

            cur_weekday = 0
            row = 4
            while row <= permanent.SCHEDULE_LAST_ROW:
                first_col_value = get_value(ws, row, 1)  # time or weekday
                # extract time
                if isinstance(first_col_value, str) and first_col_value.upper() in permanent.WEEKDAYS:
                    cur_weekday += 1
                    row += 1
                    continue

                cell_new = parse_cell(ws, row, col)
                if not cell_new[0]:
                    row += 3
                    continue
                if cell_new[0] == -1:
                    # send error notification to admins
                    for admin in SUPERADMIN_LIST:
                        bot.send_message(admin, f"{permanent.MESSAGE_ERROR_PARSE_SYNTAX} row={row} col={col}")
                    row += 3
                    continue

                time_splitted = first_col_value.split('-')
                start_time, end_time = time_splitted[0], time_splitted[1]
                subject, teacher, room = cell_new

                # usually subject is assigned to course group, but not always (e.g. B20 English groups)
                subject_group = course_group

                is_english = "English" in subject
                # extract english group number from circle brackets
                if is_english:
                    subject_group = list(map(str.strip, re.split(r"\s*/\s*", re.split(r"\s*-\s*", subject)[1])))
                    subject = re.split(r"\s*-\s*", subject)[0]
                    room1 = re.split(r"\s*/\s*", room if isinstance(room, str) else "")
                    teacher1 = re.split(r"\s*/\s*", teacher if isinstance(room, str) else "")
                    if len(room1) == len(subject_group):
                        room = room1
                    else:
                        room = [room] * len(subject_group)
                    if len(teacher1) == len(subject_group):
                        teacher = teacher1
                    else:
                        teacher = [teacher] * len(subject_group)

                lesson_id = (subject_group, subject, teacher, room, start_time, end_time, cur_weekday)
                if lesson_id in memorized_lessons:
                    row += 3
                    continue
                memorized_lessons.append(lesson_id)

                if compare_with_prev:
                    # compare new cell with old one
                    cell_old = parse_cell(ws_old, row, col)
                    if cell_new != cell_old:
                        subject_old, teacher_old, room_old = cell_old[0], cell_old[1], cell_old[2]
                        for admin in SUPERADMIN_LIST:
                            # send changes to admin
                            print(f"{subject_group} {first_col_value} changed:\n"
                                  f"Was {subject_old}, {teacher_old}, {room_old}\n"
                                  f"Now {subject}, {teacher}, {room}\n")
                            # bot.send_message(admin, f"{subject_group} {first_col_value} changed:\n"
                            #                         f"Was {subject_old}, {teacher_old}, {room_old}\n"
                            #                         f"Now {subject}, {teacher}, {room}\n")

                # insert new lesson to database
                if is_english:
                    for group, room, teacher in zip(subject_group, room, teacher):
                        controller.insert_lesson(group, subject, teacher, cur_weekday, start_time, end_time,
                                                 room)
                else:
                    controller.insert_lesson(subject_group, subject, teacher, cur_weekday, start_time, end_time, room)

                row += 3
            col += 1

        # add special lessons here manually
        # controller.insert_lesson("B17-03", "SQL injections", "Nikolai Mikriukov", 0, "13:37", "15:00", 108)

    # open parse function to other modules
    attach_autoparser_module.parse_schedule_func = parse_new_timetable
    # add parse function call to schedule on each day
    schedule.every().day.at(permanent.ADMIN_NOTIFY_TIME).do(parse_new_timetable)
